# Run this app with `python app.py` and
# visit http://127.0.0.1:8050/ in your web browser.
import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State
import pandas as pd
from datetime import datetime
import datetime
import base64
import locale
import graficos
import zipfile
import io
import matplotlib.pyplot as plt
from plotly.offline import plot
import numpy as np
import plotly.offline as offline
import plotly.graph_objs as go
import plotly.express as px
import tempfile
import os
import calendar
import graficos
import openpyxl
from openpyxl import workbook 
from openpyxl import load_workbook
from flask import Flask, send_file, make_response
import requests
import colorlover as cl

def mensal_bar(mes,tipo,ano,link,df): #ex: 1,Presencial
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	if tipo   != 'todos':
		df     = df[df['TIPO']== tipo]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['GRUPO','SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
		gray_palette = cl.scales['9']['seq']['Greys']
		fig = px.bar(df_sum, x='Hora', y='GRUPO', color='SUBCATEGORIA', orientation='h',color_discrete_sequence=px.colors.sequential.Greys)
		fig.update_layout(title='Horas total trabalhadas por categoria e subcategoria')
		fig.update_layout( xaxis_title='Horas',yaxis_title='Categoria',legend_title='Subcategoria')
		fig.update_traces(marker=dict(line=dict(width=1, color='black')))
		if link == 1:
			fig=offline.plot(fig,output_type='div')
		return fig
	
def diario_bar (dia,tipo,df): #ex: 10/01/2022,Remoto
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	dia        = datetime.datetime.strptime(dia, '%d/%m/%Y')
	df         = df[df['DATA']== dia]
	if tipo   != 'todos':
		df     = df[df['TIPO']== tipo]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['GRUPO','SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
		fig = px.bar(df_sum, x='Hora', y='GRUPO', color='SUBCATEGORIA', orientation='h',color_discrete_sequence=px.colors.sequential.Greys)
		fig.update_layout( xaxis_title='Horas',yaxis_title='Categoria',legend_title='Subcategoria')
		fig.update_layout(title='Horas total trabalhadas no dia por categoria e subcategoria')
		fig.update_traces(marker=dict(line=dict(width=1, color='black')))
		return fig
	
def mensal_line(mes,tipo,ano,link,df): #ex: 1,Remoto
	year       = ano
	month      = mes
	start_date = pd.Timestamp(year, month, 1)
	end_date   = start_date + pd.offsets.MonthEnd(0)
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	if tipo   != 'todos':
		df     = df[df['TIPO']== tipo]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['GRUPO','DATA']).agg({'Hora': 'sum'}).reset_index()
		fig = px.bar(df_sum, x='DATA', y='Hora', color='GRUPO', orientation='v',color_discrete_sequence=['Black','Grey'])
		fig.update_layout(xaxis_title='Data', yaxis_title='Horas', legend_title='Grupo')
		fig.update_layout(xaxis_range=[start_date,end_date])
		fig.update_layout(title='Horas trabalhadas por dia')
		fig.update_layout(xaxis_tickmode='linear')
		fig.update_layout(xaxis_tickangle=-90)
		if link ==1:
			fig=offline.plot(fig,output_type='div')
		return fig

def mensal_todos(mes,ano,link,df): #ex: 1
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	if len(df)==0:
		return 'nan'
	else:
		df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
		df_sum     = df.groupby(['TIPO']).agg({'Hora': 'sum'}).reset_index()
		fig        = px.bar(df_sum, x='TIPO', y='Hora', color='TIPO', orientation='v',color_discrete_sequence=['Black','Grey'])
		fig.update_layout(yaxis_title='Horas',xaxis_title=' ',legend_title='Tipo')
		fig.update_layout(title='Horas total trabalhadas por tipo')
		if link ==1:
			fig=offline.plot(fig,output_type='div')
		return fig
        
def convert_to_time(decimal_num):
    hours = int(decimal_num)
    minutes = int(round((decimal_num - hours) * 60))
    return f"{hours:02d}:{minutes:02d}"

def preenche_modelo(mes,ano,nome,df): #ex: 1,Presencial
	url = 'https://github.com/Grenda07/ppgtu/blob/main/modelo.xlsx?raw=true'
	response = requests.get(url)
	content = response.content
	file = io.BytesIO(content)
	grupos=['Grupo de Pesquisa','Programa']
	tipo=['Presencial','Remoto']
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
	month_name = datetime.date(2000, mes, 1).strftime('%B')
	month_name = month_name.capitalize()
	meses = {"January": "Janeiro",
	         "February": "Fevereiro",
	         "March": "Março",
	         "April": "Abril",
	         "May": "Maio",
	         "June": "Junho",
	         "July": "Julho",
	         "August": "Agosto",
	         "September": "Setembro",
	         "October": "Outubro",
	         "November": "Novembro",
	         "December": "Dezembro"}
	wb = load_workbook(file)
	sheets = wb.sheetnames
	Sheet1 = wb[sheets[0]] ##
	#AQUI
	Sheet1.cell(row = 3, column = 1).value = nome
	Sheet1.cell(row = 3, column = 3).value = meses[month_name]
	Sheet1.cell(row = 3, column = 4).value = ano
	total= df['Hora'].sum() 
	Sheet1.cell(row = 10, column = 3).value = convert_to_time(total)
	Sheet1.cell(row = 7, column = 3).value = convert_to_time(total)
	df_sum     = df.groupby(['GRUPO']).agg({'Hora': 'sum'}).reset_index()
	df_sum2    = df.groupby(['TIPO']).agg({'Hora': 'sum'}).reset_index()
	for index, (n, i) in enumerate(zip(grupos, tipo)):
		horas = df_sum.loc[df_sum['GRUPO'] == n, 'Hora'].iloc[0]
		Sheet1.cell(row = 8 + index, column = 3).value = convert_to_time(horas)
		Sheet1.cell(row = 8 + index, column = 4).value = horas*100/total
		horas = df_sum2.loc[df_sum2['TIPO'] == i, 'Hora'].iloc[0]
		Sheet1.cell(row = 5 + index, column = 3).value = convert_to_time(horas)
		Sheet1.cell(row = 5 + index, column = 4).value = horas*100/total
	df_sum3    = df.groupby(['SUBCATEGORIA','GRUPO']).agg({'Hora': 'sum'}).reset_index()
	for index, row in df_sum3.iterrows():
		Sheet1.cell(row = 12 + index, column = 1).value = row['SUBCATEGORIA']
		Sheet1.cell(row = 12 + index, column = 4).value = row['GRUPO']
		Sheet1.cell(row = 12 + index, column = 2).value = convert_to_time(row['Hora'])
		df4    = df[df['SUBCATEGORIA']==row['SUBCATEGORIA']]
		Sheet1.cell(row = 12 + index, column = 3).value = df4['ATIVIDADE'].nunique()
	Sheet1.title = 'Relatorio'
	df = df.drop(['Hora'], axis=1)
	df['DATA'] = df['DATA'].dt.strftime('%d/%m/%Y') 
	df['HORAS'] = df['HORAS'].dt.strftime('%H:%M')
	
	df.to_excel('temp.xlsx', sheet_name='Dados detalhados',index=False)
	temp_wb = openpyxl.load_workbook('temp.xlsx')
	ws2 = wb.create_sheet("Dados detalhados")
	for row in temp_wb['Dados detalhados']:
		for cell in row:
			ws2[cell.coordinate].value = cell.value
	new_file_name = 'relatorio_'+meses[month_name]+'.xlsx'
	wb.save(new_file_name)
	return new_file_name
	
def retorna_df(contents, filename):
	contents=str(contents[0])
	content_type, content_string = contents.split(',')
	decoded = base64.b64decode(content_string)
	df = pd.read_excel(io.BytesIO(decoded))
	return df


# Initialize the app
app = dash.Dash(__name__)
server=app.server

# Define the layout
app.layout = html.Div([
    html.H1('Relatório Atividades PPGTU',style={'textAlign': 'center','fontFamily':'Arial','fontSize':34}),
    html.Div([dcc.Upload(id='upload-data',children=html.Div(['Arraste e solte ou ',html.A('selecione arquivos')]),
              style={'width': '600px','height': '60px','lineHeight': '60px','borderWidth': '1px',
                     'borderStyle': 'dashed','borderRadius': '5px','textAlign': 'center','fontFamily':'Arial','fontSize':14},
              multiple=True),
              html.Button('Submeter', id='transform-button',style={'position': 'relative','textAlign': 'center','fontFamily':'Arial','fontSize':14}),
              html.Label(id='output-data-upload', style={'display': 'inline-block','fontFamily':'Arial','fontSize':14})],
              style={'display': 'flex', 'alignItems': 'center', 'justifyContent': 'center', 'flexDirection': 'column'}),
    html.Div(id='tipo-container', children=[
    dcc.RadioItems(
        id='freq-tipo',
        options=[
            {'label': 'Presencial', 'value': 'Presencial'},
            {'label': 'Remoto', 'value': 'Remoto'},
            {'label': 'Presencial e Remoto', 'value': 'todos'},
            {'label': 'Relatório' , 'value': 'relatorio'}
        ], value=None
    )],style={'textAlign': 'center','fontFamily':'Arial','fontSize':14,'flex': 1}),
    
    html.Div(id='rela-container', children=[html.Br(),
    dcc.RadioItems(
        id='freq-radio',
        options=[
            {'label': 'Mensal', 'value': 'mensal'},
            {'label': 'Diário', 'value': 'diario'}
        ], value=None,style={'textAlign': 'center','fontFamily':'Arial','fontSize':14,'flex': 1}
    ),html.Br()], style={'display': 'none'}),  
    
    html.Div(id='mensal-container',children=[html.Div([html.Div([html.Label('Ano:'),html.Br(),
                                                                 dcc.Input(id='year-input',type='number',placeholder='Ano',style={'textAlign': 'center','width': '200px','height':'25px'})],
                                                                 style={'display': 'inline-block','margin-right': '20px'}),
                                                       html.Div([html.Label('Mês:'),html.Br(),dcc.Dropdown(id='month-dropdown',
                            options=[
                                {'label': 'Janeiro', 'value': '01'},
                                {'label': 'Fevereiro', 'value': '02'},
                                {'label': 'Março', 'value': '03'},
                                {'label': 'Abril', 'value': '04'},
                                {'label': 'Maio', 'value': '05'},
                                {'label': 'Junho', 'value': '06'},
                                {'label': 'Julho', 'value': '07'},
                                {'label': 'Agosto', 'value': '08'},
                                {'label': 'Setembro', 'value': '09'},
                                {'label': 'Outubro', 'value': '10'},
                                {'label': 'Novembro', 'value': '11'},
                                {'label': 'Dezembro', 'value': '12'},
                            ],
                            placeholder='Mês',
                            style={'width': '200px','height':'10px'})],
                    style={
                        'display': 'inline-block',
                        'margin-right': '10px','vertical-align': 'top'})],
            style={
                'text-align': 'center',
                'font-family': 'Arial',
                'font-size': '14px',
            },
        ),
        html.Br(),
        html.Button('Enter', id='submit-btn',style={'display': 'block', 'margin': 'auto'}),
    ],
    style={'display': 'none'}),
    
    html.Div(children=[html.Div(id='diario-container',
                                children=[dcc.Input(id='date-input',type='text',placeholder='DD/MM/YYYY'),
                                                    html.Button('Enter', id='submit-btn-2')],
                                style={'display': 'none','textAlign': 'center',
                                       'fontFamily': 'Arial','fontSize': 14,
                                       'width': '400px','margin': '0 auto', 'marginBottom': '20px'})],
             style={'display': 'flex', 'justifyContent': 'center'}),
             
    
    
    html.Div(id='caixa-error',children=[html.Br(),html.Div(id='mensagem', style={'display': 'inline-block', 'marginLeft': '450px',
                                                              'fontFamily':'Arial','fontSize':20}),html.Br()],
             style={'display': 'none'}),
    
    
             
    html.Div(id='caixa-error2',children=[html.Br(),html.Div(id='mensagem2', style={'display': 'inline-block', 'marginLeft': '450px',
                                                              'fontFamily':'Arial','fontSize':20})],
             style={'display': 'none'}),
    
    html.Div(id="mensal-graphs1", children=[
        dcc.Graph(id="graph-1-mes"),
        dcc.Graph(id="graph-2")
    ], style={'display': 'none'}),
    
    html.Div(id="mensal-graphs2", children=[
        dcc.Graph(id="graph-1-1-mes"),
        dcc.Graph(id="graph-2-2"),
        dcc.Graph(id="graph-3")
    ], style={'display': 'none'}),
    
    html.Div(id="diario-graphs", children=[
        dcc.Graph(id="graph-1-dia")
    ], style={'display': 'none'}),
    
             
    html.Div(id='caixa-error3',children=[html.Br(),html.Div(id='mensagem3', style={'display': 'inline-block', 'marginLeft': '450px',
                                                              'fontFamily':'Arial','fontSize':20})],
             style={'display': 'none'}),
	html.Div(id="relatorio-container",style={'display': 'none'}, children=[html.Br(),
        html.Div( children=[ 
        html.Label('Ano:'), html.Br(),
        dcc.Input(id='year-input2', type='number', placeholder='Ano',style={'width': '450px'}), html.Br(),html.Br(),
        html.Label('Mês:'),html.Br(),
        dcc.Dropdown(
            id='month-dropdown2',
            options=[
                {'label': 'Janeiro'  , 'value': '01'},
                {'label': 'Fevereiro', 'value': '02'},
                {'label': 'Março'    , 'value': '03'},
                {'label': 'Abril'    , 'value': '04'},
                {'label': 'Maio'     , 'value': '05'},
                {'label': 'Junho'    , 'value': '06'},
                {'label': 'Julho'    , 'value': '07'},
                {'label': 'Agosto'   , 'value': '08'},
                {'label': 'Setembro' , 'value': '09'},
                {'label': 'Outubro'  , 'value': '10'},
                {'label': 'Novembro' , 'value': '11'},
                {'label': 'Dezembro' , 'value': '12'}
            ],
            placeholder='Mês',style={'width': '450px'}
        ),html.Br(),html.Label('Aluno(a):'),html.Br(),
        dcc.Input(id='name-input', type='text', placeholder='Digite seu nome',style={'width': '450px'}),
        html.Br(),html.Br(),html.Button('Gerar Relatório', id='download-link'),
        dcc.Download(id='download')
    ], style={'border': '2px solid black', 'padding': '20px', 'textAlign': 'left', 'maxWidth': '500px',"margin": "0 auto"}),
    html.Br()
])

])


# Define the callbacks
@app.callback(
    [Output('output-data-upload', 'children'),Output('mensal-container', 'style'), 
     Output('diario-container', 'style')     ,Output("mensal-graphs1", "style")  ,
     Output("mensal-graphs2", "style")       ,Output("diario-graphs", "style")   ,
     Output("relatorio-container", "style")  ,Output("rela-container", "style")  ,
     Output("caixa-error", "style")          ,Output("caixa-error2", "style"),
     Output("caixa-error3", "style")],
    [Input('freq-radio', 'value'),Input('freq-tipo', 'value'),Input('transform-button', 'n_clicks')],
    [State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def show_hide_divs(frequency,tipo,n_clicks,contents,filename):
	if not n_clicks or not contents or not filename:
		mensal_style         = {'display': 'none'}
		diario_style         = {'display': 'none'}
		mensal_graphs1_style = {'display': 'none'}
		mensal_graphs2_style = {'display': 'none'}
		diario_graphs_style  = {'display': 'none'}
		gerar_style          = {'display': 'none'}
		tipo_style           = {'display': 'none'}
		caixa_style          = {'display': 'none'}
		caixa2_style         = {'display': 'none'}
		caixa3_style         = {'display': 'none'}
		return {},mensal_style, diario_style, mensal_graphs1_style, mensal_graphs2_style, diario_graphs_style,gerar_style,tipo_style,caixa_style,caixa2_style,caixa3_style
	mensal_style         = {'display': 'block'} if tipo      != 'relatorio' and frequency == "mensal" else {'display': 'none'}
	diario_style         = {'display': 'block'} if tipo      != 'relatorio' and frequency == "diario" else {'display': 'none'}
	mensal_graphs1_style = {'display': 'block'} if tipo      != 'relatorio' and frequency == "mensal" and tipo != 'todos' else {'display': 'none'}
	mensal_graphs2_style = {'display': 'block'} if tipo      != 'relatorio' and frequency == "mensal" and tipo == 'todos' else {'display': 'none'}
	diario_graphs_style  = {'display': 'block'} if tipo      != 'relatorio' and frequency == "diario" else {'display': 'none'}
	gerar_style          = {'display': 'block'} if tipo      == 'relatorio' else {'display': 'none'}
	tipo_style           = {'display': 'block'} if tipo      == 'todos' or tipo == 'Remoto' or tipo == 'Presencial' else {'display': 'none'}
	caixa_style          = {'display': 'block'} if frequency == "mensal" and tipo != 'relatorio' else {'display': 'none'}
	caixa2_style         = {'display': 'block'} if frequency == "diario" and tipo != 'relatorio' else {'display': 'none'}
	caixa3_style         = {'display': 'block'} if tipo      == 'relatorio' else {'display': 'none'}
	return html.Div('Ok!'),mensal_style, diario_style, mensal_graphs1_style, mensal_graphs2_style, diario_graphs_style,gerar_style,tipo_style,caixa_style,caixa2_style,caixa3_style

## PARA MENSAL (sem todos):
@app.callback(
    [Output('mensagem', 'children'),Output('graph-1-mes', 'figure'), Output('graph-2', 'figure')],
    [Input('submit-btn', 'n_clicks'),Input('freq-tipo', 'value')],
    [State('year-input', 'value'), State('month-dropdown', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_1(n_clicks,tipo, year, month,contents, filename):
	if not year or not month:
		return {},{}, {}
	df = retorna_df(contents, filename)
	fig1=mensal_bar(int(month),tipo,year,0,df)
	fig2=mensal_line(int(month),tipo,year,0,df)
	
	if fig1 == 'nan' or fig2 == 'nan':
		return html.Div('Não há dados para esse período.'),{},{}
	else:
		return {},fig1, fig2

#PARA MENSAL (com todos):
@app.callback(
    [Output('graph-1-1-mes', 'figure'), Output('graph-2-2', 'figure'), Output('graph-3', 'figure')],
    [Input('submit-btn', 'n_clicks'),Input('freq-tipo', 'value')],
    [State('year-input', 'value'), State('month-dropdown', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_2(n_clicks,tipo, year, month,contents, filename):
	if not year or not month:
		return {}, {},{}
	df = retorna_df(contents, filename)
	fig1=mensal_bar(int(month),tipo,year,0,df)
	fig2=mensal_line(int(month),tipo,year,0,df)
	fig3=mensal_todos(int(month),year,0,df)
	if fig1 == 'nan' or fig2 == 'nan' or fig3 == 'nan':
		return {},{},{}
	else:
		return fig1, fig2,fig3

#PARA RELATÓRIO:
@app.callback(
    [Output('download', 'data'),Output('mensagem3', 'children')],
    [Input('download-link', 'n_clicks'),Input('name-input', 'value')],
    [State('year-input2', 'value'), State('month-dropdown2', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_2(n_clicks,nome, year, month,contents, filename):
	if not year or not month:
		return None,{}
	df  = retorna_df(contents, filename)
	df = df.loc[df['DATA'].dt.month == int(month)]
	df = df.loc[df['DATA'].dt.year == year]
	if n_clicks is not None:
		if len(df)==0:
			return None,html.Div('Não há dados para esse período.')
		return dcc.send_file(preenche_modelo(int(month),year,nome,df)),{}
	else:
		return None,{}

#PARA DIÁRIO:
@app.callback(
    [Output('mensagem2', 'children'),Output('graph-1-dia', 'figure')],
    [Input('submit-btn-2', 'n_clicks'),Input('freq-tipo', 'value')],
    [State('date-input', 'value'),State('upload-data', 'contents'),State('upload-data', 'filename')]
)
def update_graphs_3(n_clicks,tipo, date,contents, filename):
	if not date:
		return {},{}
	df = retorna_df(contents, filename)
	fig = diario_bar (date,tipo,df)
	if fig == 'nan':
		return html.Div('Não há dados para esse período.'),{}
	else:
		return {},fig

if __name__ == '__main__':
    app.run_server(debug=True)
